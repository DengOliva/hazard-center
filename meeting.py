import json
import os
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path

from flask import Blueprint, jsonify, request, send_from_directory

ROOT = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", ROOT / "data"))
DB_PATH = DATA_DIR / "hazards.db"
MEETING_DATA_FILE = DATA_DIR / "meeting_data.json"

bp = Blueprint("meeting", __name__)


@bp.get("/meeting")
def meeting_page():
    return send_from_directory(ROOT / "public" / "meeting", "index.html")


def _db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _week_range(date_str=None):
    """Return (start, end) for the ISO week containing date_str or today."""
    if not date_str:
        date_str = date.today().isoformat()
    d = date.fromisoformat(date_str)
    start = d - timedelta(days=d.weekday())
    end = start + timedelta(days=6)
    return start.isoformat(), end.isoformat()


def _load_alert_data():
    """Load alert data — reuse the same logic as app.py's load_alert_data."""
    import re as _re
    from openpyxl import load_workbook as _load_wb

    ALERT_DATA_FILE = DATA_DIR / "alert_台账.xlsx"
    ALERT_SEED_FILE = ROOT / "seed" / "01 防城港三期安全管理数据总台账.xlsx"

    src = ALERT_DATA_FILE if ALERT_DATA_FILE.exists() else (ALERT_SEED_FILE if ALERT_SEED_FILE.exists() else None)
    if not src:
        return None

    wb = _load_wb(src, data_only=True)

    MONTHS = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]

    def _iso_date(value):
        if value is None or value == "":
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (int, float)):
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text[:19], fmt).date().isoformat()
            except ValueError:
                pass
        return text[:10]

    DETAIL_SHEET_CONFIG = [
        ("工程公司挂牌督办单", "external", "挂牌督办", 2, 8, 11),
        ("红黄牌", "external", "红黄牌", 3, 5, 12),
        ("工程公司处理通报", "external", "处理通报", 4, 8, 10),
        ("工程公司通报批评", "external", "处理通报", 5, 0, 11),
        ("工程公司整改单", "external", "整改单", 6, 9, 12),
        ("工程公司停工令", "external", "停工令", 5, 10, 13),
        ("监理业主整改通知单", "external", "监理通知单", 5, 7, 10),
        ("工程公司违章培训通知单", "external", "违章培训通知单", 7, 12, 3),
        ("项目内部处理通报", "internal", "处理通报", 3, 7, 8),
        ("项目整改通知单", "internal", "整改单", 3, 7, 8),
        ("项目停工令", "internal", "停工令", 2, 6, 10),
        ("项目违章培训通知单", "internal", "违章培训通知单", 7, 10, 4),
    ]

    detail_records = []
    for sheet_name, category, type_name, date_col, dept_col, sub_col in DETAIL_SHEET_CONFIG:
        if sheet_name not in wb.sheetnames:
            continue
        dws = wb[sheet_name]
        for r in range(2, dws.max_row + 1):
            date_val = dws.cell(row=r, column=date_col).value
            if date_val is None:
                continue
            parsed = _iso_date(date_val)
            if not parsed or len(parsed) < 10:
                continue
            dept_name = ""
            if dept_col > 0:
                dept_name = str(dws.cell(row=r, column=dept_col).value or "").strip()
                if dept_name in ("/", "None", "#N/A", ""):
                    dept_name = ""
            sub_name = ""
            if sub_col > 0:
                sub_name = str(dws.cell(row=r, column=sub_col).value or "").strip()
                if sub_name in ("/", "None", "#N/A", ""):
                    sub_name = ""
            detail_records.append({
                "date": parsed, "dept_name": dept_name, "sub_name": sub_name,
                "category": category, "type_name": type_name, "sheet_name": sheet_name,
            })

    wb.close()
    return detail_records


@bp.get("/api/meeting/data")
def meeting_data():
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    if not start or not end:
        start, end = _week_range()

    result = {"start": start, "end": end}

    # --- Slide 7: 内外部行为偏差统计 (from alert detail records) ---
    alert_records = _load_alert_data()
    behavior_stats = {"internal": [], "external": []}
    if alert_records:
        type_map_internal = {}
        type_map_external = {}
        for rec in alert_records:
            if rec["date"] < start or rec["date"] > end:
                continue
            tn = rec["type_name"]
            if rec["category"] == "internal":
                type_map_internal[tn] = type_map_internal.get(tn, 0) + 1
            else:
                type_map_external[tn] = type_map_external.get(tn, 0) + 1
        behavior_stats["internal"] = [{"type": k, "count": v} for k, v in type_map_internal.items()]
        behavior_stats["external"] = [{"type": k, "count": v} for k, v in type_map_external.items()]

    result["behavior_stats"] = behavior_stats

    # --- Slide 8: 隐患记录提交 (people statistics) ---
    with _db() as conn:
        people_rows = conn.execute(
            "SELECT * FROM people WHERE active=1 ORDER BY category, department, name"
        ).fetchall()
        counts = {
            r[0]: r[1] for r in conn.execute(
                "SELECT checker_name, COUNT(*) FROM hazards WHERE check_date BETWEEN ? AND ? GROUP BY checker_name",
                (start, end),
            )
        }
        b_counts = {
            r[0]: r[1] for r in conn.execute(
                "SELECT checker_name, COUNT(*) FROM hazards WHERE check_date BETWEEN ? AND ? AND hazard_level='B' GROUP BY checker_name",
                (start, end),
            )
        }

    people_stats = []
    for p in people_rows:
        count = counts.get(p["name"], 0)
        people_stats.append({
            "name": p["name"], "category": p["category"],
            "department": p["department"], "count": count,
            "b_count": b_counts.get(p["name"], 0),
        })

    # Group by category and department
    dept_order = ["安监部", "质控部", "技术部", "工程部", "物资部", "机械设备管理部", "经理部"]
    groups = {}
    for p in people_stats:
        key = p["category"]
        if key not in groups:
            groups[key] = {
                "category": key,
                "depts": {},
                "members": [],
                "total": 0,
            }
        groups[key]["members"].append(p)
        groups[key]["total"] += p["count"]
        dept = p["department"] or "其他"
        if dept not in groups[key]["depts"]:
            groups[key]["depts"][dept] = 0
        groups[key]["depts"][dept] += p["count"]

    result["people_stats"] = {
        "groups": list(groups.values()),
        "total_overall": sum(g["total"] for g in groups.values()),
        "dept_order": dept_order,
    }

    # --- Slide 11: 管理干部预警刹车统计 (from alert data) ---
    if alert_records:
        dept_alert = {}
        ext_types = ["挂牌督办", "红黄牌", "处理通报", "整改单", "停工令", "违章培训通知单", "监理通知单"]
        int_types = ["停工令", "处理通报", "整改单", "违章培训通知单"]
        for rec in alert_records:
            if rec["date"] < start or rec["date"] > end:
                continue
            if not rec["dept_name"]:
                continue
            dept = rec["dept_name"]
            if dept not in dept_alert:
                dept_alert[dept] = {"dept": dept}
                for t in ext_types:
                    dept_alert[dept][f"ext_{t}"] = 0
                for t in int_types:
                    dept_alert[dept][f"int_{t}"] = 0
            key = f"ext_{rec['type_name']}" if rec["category"] == "external" else f"int_{rec['type_name']}"
            if key in dept_alert[dept]:
                dept_alert[dept][key] += 1
        result["dept_alert"] = {
            "ext_types": ext_types,
            "int_types": int_types,
            "depts": list(dept_alert.values()),
        }
    else:
        result["dept_alert"] = {"ext_types": [], "int_types": [], "depts": []}

    # --- Slides 12-14: 整改单/通报/停工令 close status ---
    result["close_status"] = _build_close_status(alert_records, start, end)

    # --- Slide 16: 培训 ---
    result["training"] = _build_training_summary(start, end)

    return jsonify(result)


def _build_close_status(alert_records, start, end):
    """Parse close status from alert detail sheets for 整改单/通报/停工令."""
    from openpyxl import load_workbook as _load_wb

    ALERT_DATA_FILE = DATA_DIR / "alert_台账.xlsx"
    ALERT_SEED_FILE = ROOT / "seed" / "01 防城港三期安全管理数据总台账.xlsx"
    src = ALERT_DATA_FILE if ALERT_DATA_FILE.exists() else (ALERT_SEED_FILE if ALERT_SEED_FILE.exists() else None)
    if not src:
        return {"rectification": [], "notice": [], "stop_work": []}

    wb = _load_wb(src, data_only=True)

    def _iso(v):
        if v is None or v == "":
            return ""
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        if isinstance(v, (int, float)):
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date().isoformat()
        text = str(v).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text[:19], fmt).date().isoformat()
            except ValueError:
                pass
        return text[:10]

    def _parse_close_sheet(sheet_name, parse_rows=None):
        records = []
        if sheet_name not in wb.sheetnames:
            return records
        ws = wb[sheet_name]
        headers = []
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            headers.append(str(h).strip() if h else "")

        max_row = min(ws.max_row + 1, parse_rows) if parse_rows else ws.max_row + 1
        for r in range(2, max_row):
            row_vals = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                header = headers[c - 1] if c <= len(headers) else f"col_{c}"
                row_vals[header] = str(v).strip() if v is not None else ""
            if all(v == "" for v in row_vals.values()):
                continue
            records.append(row_vals)
        return records

    result = {"rectification": [], "notice": [], "stop_work": []}

    # 工程公司整改单
    for rec in _parse_close_sheet("工程公司整改单"):
        result["rectification"].append({
            "source": "工程公司", "no": rec.get("整改单编号", rec.get("整改通知单编号", "")),
            "dept": rec.get("责任部门", rec.get("检查部门", "")),
            "issue_date": _iso(rec.get("检查日期", rec.get("下发日期", ""))),
            "content": rec.get("整改内容", rec.get("问题描述", ""))[:60],
            "deadline": _iso(rec.get("整改期限", rec.get("要求完成日期", ""))),
            "closed": rec.get("是否关闭", rec.get("是否整改", "")),
            "remark": rec.get("关闭情况说明", rec.get("备注", ""))[:40],
        })
    # 项目整改通知单
    for rec in _parse_close_sheet("项目整改通知单"):
        result["rectification"].append({
            "source": "项目部", "no": rec.get("整改通知单编号", rec.get("整改单编号", "")),
            "dept": rec.get("责任部门", rec.get("检查部门", "")),
            "issue_date": _iso(rec.get("检查日期", rec.get("下发日期", ""))),
            "content": rec.get("整改内容", rec.get("问题描述", ""))[:60],
            "deadline": _iso(rec.get("整改期限", rec.get("要求完成日期", ""))),
            "closed": rec.get("是否关闭", rec.get("是否整改", "")),
            "remark": rec.get("关闭情况说明", rec.get("备注", ""))[:40],
        })

    # 工程公司处理通报
    for rec in _parse_close_sheet("工程公司处理通报"):
        result["notice"].append({
            "source": "工程公司", "no": rec.get("通报编号", rec.get("处理通报编号", "")),
            "dept": rec.get("责任部门", rec.get("检查部门", "")),
            "issue_date": _iso(rec.get("检查日期", rec.get("下发日期", ""))),
            "content": rec.get("通报内容", rec.get("问题描述", ""))[:60],
            "deadline": _iso(rec.get("整改期限", rec.get("要求完成日期", ""))),
            "closed": rec.get("是否关闭", rec.get("是否整改", "")),
        })
    # 项目内部处理通报
    for rec in _parse_close_sheet("项目内部处理通报"):
        result["notice"].append({
            "source": "项目部", "no": rec.get("通报编号", rec.get("处理通报编号", "")),
            "dept": rec.get("责任部门", rec.get("检查部门", "")),
            "issue_date": _iso(rec.get("检查日期", rec.get("下发日期", ""))),
            "content": rec.get("通报内容", rec.get("问题描述", ""))[:60],
            "deadline": _iso(rec.get("整改期限", rec.get("要求完成日期", ""))),
            "closed": rec.get("是否关闭", rec.get("是否整改", "")),
        })

    # 工程公司停工令
    for rec in _parse_close_sheet("工程公司停工令"):
        result["stop_work"].append({
            "source": "工程公司", "no": rec.get("停工令编号", ""),
            "dept": rec.get("责任部门", rec.get("检查部门", "")),
            "issue_date": _iso(rec.get("检查日期", rec.get("下发日期", ""))),
            "content": rec.get("停工内容", rec.get("问题描述", ""))[:60],
            "deadline": _iso(rec.get("整改期限", rec.get("要求完成日期", ""))),
            "closed": rec.get("是否关闭", rec.get("是否整改", "")),
        })
    # 项目停工令
    for rec in _parse_close_sheet("项目停工令"):
        result["stop_work"].append({
            "source": "项目部", "no": rec.get("停工令编号", ""),
            "dept": rec.get("责任部门", rec.get("检查部门", "")),
            "issue_date": _iso(rec.get("检查日期", rec.get("下发日期", ""))),
            "content": rec.get("停工内容", rec.get("问题描述", ""))[:60],
            "deadline": _iso(rec.get("整改期限", rec.get("要求完成日期", ""))),
            "closed": rec.get("是否关闭", rec.get("是否整改", "")),
        })

    wb.close()
    return result


def _build_training_summary(start, end):
    """Build training summary from training overrides/materials."""
    TRAINING_SCHEDULE_FILE = ROOT / "seed" / "2026年7月安全培训安排表.xlsx"
    result = {"items": [], "materials_count": 0}

    if TRAINING_SCHEDULE_FILE.exists():
        from openpyxl import load_workbook as _load_wb
        import re as _re

        wb = _load_wb(TRAINING_SCHEDULE_FILE, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        for row in sheet.iter_rows(min_row=3, values_only=True):
            schedule_date = ""
            if len(row) > 0:
                text = str(row[0] or "").strip()
                match = _re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", text)
                if match:
                    y, m, d = map(int, match.groups())
                    schedule_date = date(y, m, d).isoformat()
            if not schedule_date or schedule_date < start or schedule_date > end:
                continue
            morning = str(row[1] or "").strip() if len(row) > 1 else ""
            afternoon = "、".join(str(row[i] or "").strip() for i in (3, 4) if len(row) > i and row[i])
            night = str(row[5] or "").strip() if len(row) > 5 else ""
            for text in [morning, afternoon, night]:
                if text:
                    items = [it.strip() for it in _re.split(r"[\n\r]+", text) if it.strip()]
                    for item in items:
                        result["items"].append({
                            "date": schedule_date,
                            "content": item,
                        })
        wb.close()

    TRAINING_MATERIALS_FILE = DATA_DIR / "training_materials.json"
    if TRAINING_MATERIALS_FILE.exists():
        try:
            materials = json.loads(TRAINING_MATERIALS_FILE.read_text(encoding="utf-8"))
            result["materials_count"] = len(materials)
        except Exception:
            pass

    return result


@bp.get("/api/meeting/manual")
def meeting_manual_load():
    """Load saved manual meeting data."""
    if not MEETING_DATA_FILE.exists():
        return jsonify({})
    try:
        data = json.loads(MEETING_DATA_FILE.read_text(encoding="utf-8"))
        return jsonify(data)
    except Exception:
        return jsonify({})


@bp.post("/api/meeting/manual")
def meeting_manual_save():
    """Save manual meeting data for a specific slide."""
    body = request.get_json(force=True)
    slide_id = str(body.get("slide_id", "")).strip()
    if not slide_id:
        return jsonify(error="缺少 slide_id"), 400

    existing = {}
    if MEETING_DATA_FILE.exists():
        try:
            existing = json.loads(MEETING_DATA_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass

    existing[slide_id] = body.get("data", {})
    MEETING_DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    MEETING_DATA_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return jsonify(ok=True)
