import io
import json
import os
import shutil
import sqlite3
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

from flask import Blueprint, abort, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", ROOT / "data"))
DB_PATH = DATA_DIR / "hazards.db"
UPLOAD_DIR = DATA_DIR / "uploads" / "collections"
GENERATED_DIR = DATA_DIR / "collections"
TRAINING_EDIT_PASSWORD = os.environ.get("TRAINING_EDIT_PASSWORD", "@q")

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
GENERATED_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf", ".xlsx", ".xls", ".doc", ".docx"}
MAX_FILE_SIZE = 20 * 1024 * 1024

DEPARTMENTS = [
    "水电队", "核岛一队", "综合车间", "钢结构队", "搅拌站", "BOP队", "测量队",
    "物资部", "安监部", "技术部", "综合管理部", "商务部", "机械队", "财务部",
    "质控部", "质保部", "金属试验室",
]

bp = Blueprint("collection", __name__)


def _db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _check_password(data):
    pwd = (data or {}).get("password", "")
    if not pwd or pwd != TRAINING_EDIT_PASSWORD:
        return jsonify({"error": "密码错误"}), 403
    return None


def _allowed_file(filename):
    return "." in filename and Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


# ── Page ──

@bp.get("/collection")
def collection_page():
    return send_from_directory(ROOT / "public" / "collection", "index.html")


@bp.get("/uploads/collections/<path:filepath>")
def serve_upload(filepath):
    return send_from_directory(str(UPLOAD_DIR), filepath)


# ── Password ──

@bp.post("/api/collection/verify-password")
def verify_password():
    data = request.get_json(silent=True) or {}
    if data.get("password") != TRAINING_EDIT_PASSWORD:
        return jsonify({"ok": False, "error": "密码错误"}), 403
    return jsonify({"ok": True})


# ── Catalogs (admin) ──

@bp.get("/api/collection/catalogs")
def list_catalogs():
    with _db() as conn:
        rows = conn.execute(
            "SELECT * FROM training_collections ORDER BY created_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.post("/api/collection/catalogs")
def create_catalog():
    data = request.get_json(silent=True) or {}
    err = _check_password(data)
    if err:
        return err
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "培训名称不能为空"}), 400
    with _db() as conn:
        cur = conn.execute(
            "INSERT INTO training_collections (name, description, departments, deadline, created_at) VALUES (?,?,?,?,?)",
            (
                name,
                (data.get("description") or "").strip(),
                (data.get("departments") or "").strip(),
                (data.get("deadline") or "").strip(),
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
        conn.commit()
    return jsonify({"id": cur.lastrowid})


@bp.put("/api/collection/catalogs/<int:catalog_id>")
def update_catalog(catalog_id):
    data = request.get_json(silent=True) or {}
    err = _check_password(data)
    if err:
        return err
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "培训名称不能为空"}), 400
    with _db() as conn:
        row = conn.execute("SELECT id FROM training_collections WHERE id=?", (catalog_id,)).fetchone()
        if not row:
            return jsonify({"error": "名录不存在"}), 404
        conn.execute(
            "UPDATE training_collections SET name=?, description=?, departments=?, deadline=?, active=? WHERE id=?",
            (
                name,
                (data.get("description") or "").strip(),
                (data.get("departments") or "").strip(),
                (data.get("deadline") or "").strip(),
                data.get("active", 1),
                catalog_id,
            ),
        )
        conn.commit()
    return jsonify({"ok": True})


@bp.delete("/api/collection/catalogs/<int:catalog_id>")
def delete_catalog(catalog_id):
    data = request.get_json(silent=True) or {}
    err = _check_password(data)
    if err:
        return err
    with _db() as conn:
        conn.execute("DELETE FROM training_submissions WHERE collection_id=?", (catalog_id,))
        conn.execute("DELETE FROM training_collections WHERE id=?", (catalog_id,))
        conn.commit()
    # clean up files
    coll_dir = UPLOAD_DIR / str(catalog_id)
    if coll_dir.exists():
        shutil.rmtree(coll_dir, ignore_errors=True)
    gen_dir = GENERATED_DIR / str(catalog_id)
    if gen_dir.exists():
        shutil.rmtree(gen_dir, ignore_errors=True)
    return jsonify({"ok": True})


# ── Submissions (public) ──

@bp.get("/api/collection/submissions")
def list_submissions():
    catalog_id = request.args.get("collection_id", "")
    with _db() as conn:
        if catalog_id:
            rows = conn.execute(
                "SELECT * FROM training_submissions WHERE collection_id=? ORDER BY submitted_at DESC",
                (int(catalog_id),),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM training_submissions ORDER BY submitted_at DESC"
            ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["photos"] = json.loads(d["photos"])
        except Exception:
            d["photos"] = []
        result.append(d)
    return jsonify(result)


@bp.post("/api/collection/submit")
def submit():
    collection_id = request.form.get("collection_id", "").strip()
    department = request.form.get("department", "").strip()
    notes = request.form.get("notes", "").strip()

    if not collection_id:
        return jsonify({"error": "请选择培训名录"}), 400
    if not department:
        return jsonify({"error": "请选择部门/班组"}), 400

    # validate catalog exists and is active
    with _db() as conn:
        catalog = conn.execute(
            "SELECT id, name FROM training_collections WHERE id=? AND active=1",
            (int(collection_id),),
        ).fetchone()
        if not catalog:
            return jsonify({"error": "培训名录不存在或已停用"}), 404

    # handle sign-in file
    sign_in_file = request.files.get("sign_in_file")
    sign_in_path = ""
    if sign_in_file and sign_in_file.filename and _allowed_file(sign_in_file.filename):
        sign_in_file.seek(0, 2)
        if sign_in_file.tell() > MAX_FILE_SIZE:
            return jsonify({"error": "签到单文件超过 20MB 限制"}), 400
        sign_in_file.seek(0)
        ext = Path(sign_in_file.filename).suffix.lower()
        safe_name = f"{uuid.uuid4().hex}{ext}"
        coll_dir = UPLOAD_DIR / collection_id
        coll_dir.mkdir(parents=True, exist_ok=True)
        sign_in_file.save(str(coll_dir / safe_name))
        sign_in_path = f"{collection_id}/{safe_name}"

    # handle photo files
    photo_paths = []
    photo_files = request.files.getlist("photos")
    for pf in photo_files:
        if pf and pf.filename and _allowed_file(pf.filename):
            pf.seek(0, 2)
            if pf.tell() > MAX_FILE_SIZE:
                continue
            pf.seek(0)
            ext = Path(pf.filename).suffix.lower()
            safe_name = f"{uuid.uuid4().hex}{ext}"
            coll_dir = UPLOAD_DIR / collection_id
            coll_dir.mkdir(parents=True, exist_ok=True)
            pf.save(str(coll_dir / safe_name))
            photo_paths.append(f"{collection_id}/{safe_name}")

    with _db() as conn:
        # upsert: if same collection+department exists, update it
        existing = conn.execute(
            "SELECT id FROM training_submissions WHERE collection_id=? AND department=?",
            (int(collection_id), department),
        ).fetchone()
        if existing:
            old = conn.execute(
                "SELECT sign_in_file, photos FROM training_submissions WHERE id=?",
                (existing["id"],),
            ).fetchone()
            conn.execute(
                "UPDATE training_submissions SET sign_in_file=?, photos=?, notes=?, submitted_at=? WHERE id=?",
                (
                    sign_in_path or old["sign_in_file"],
                    json.dumps(photo_paths, ensure_ascii=False) if photo_paths else old["photos"],
                    notes,
                    datetime.now().isoformat(timespec="seconds"),
                    existing["id"],
                ),
            )
        else:
            conn.execute(
                "INSERT INTO training_submissions (collection_id, department, sign_in_file, photos, notes, submitted_at) VALUES (?,?,?,?,?,?)",
                (
                    int(collection_id),
                    department,
                    sign_in_path,
                    json.dumps(photo_paths, ensure_ascii=False),
                    notes,
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
        conn.commit()

    return jsonify({"ok": True})


@bp.delete("/api/collection/submissions/<int:submission_id>")
def delete_submission(submission_id):
    data = request.get_json(silent=True) or {}
    err = _check_password(data)
    if err:
        return err
    with _db() as conn:
        row = conn.execute(
            "SELECT id, sign_in_file, photos FROM training_submissions WHERE id=?",
            (submission_id,),
        ).fetchone()
        if not row:
            return jsonify({"error": "提交记录不存在"}), 404
        # delete files
        for path_str in [row["sign_in_file"]]:
            if path_str:
                fp = UPLOAD_DIR / path_str
                if fp.exists():
                    fp.unlink(missing_ok=True)
        try:
            for path_str in json.loads(row["photos"] or "[]"):
                fp = UPLOAD_DIR / path_str
                if fp.exists():
                    fp.unlink(missing_ok=True)
        except Exception:
            pass
        conn.execute("DELETE FROM training_submissions WHERE id=?", (submission_id,))
        conn.commit()
    return jsonify({"ok": True})


# ── Generate ledger ZIP ──

def _thin_border():
    return Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


@bp.post("/api/collection/generate/<int:catalog_id>")
def generate(catalog_id):
    data = request.get_json(silent=True) or {}
    err = _check_password(data)
    if err:
        return err

    with _db() as conn:
        catalog = conn.execute(
            "SELECT * FROM training_collections WHERE id=?", (catalog_id,)
        ).fetchone()
        if not catalog:
            return jsonify({"error": "名录不存在"}), 404
        subs = conn.execute(
            "SELECT * FROM training_submissions WHERE collection_id=? ORDER BY department",
            (catalog_id,),
        ).fetchall()

    catalog_name = catalog["name"]
    gen_root = GENERATED_DIR / str(catalog_id) / catalog_name
    if gen_root.exists():
        shutil.rmtree(gen_root, ignore_errors=True)
    gen_root.mkdir(parents=True, exist_ok=True)

    # Build ledger Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "提交台账"

    header_font = Font(name="微软雅黑", bold=True, size=11)
    cell_font = Font(name="微软雅黑", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["序号", "部门/班组", "提交时间", "签到单", "照片数", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = _thin_border()

    for row_idx, sub in enumerate(subs, 2):
        photos = json.loads(sub["photos"] or "[]")
        values = [
            row_idx - 1,
            sub["department"],
            sub["submitted_at"],
            Path(sub["sign_in_file"]).name if sub["sign_in_file"] else "",
            len(photos),
            sub["notes"] or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = center if col_idx != 6 else left_align
            cell.border = _thin_border()

    # Summary row
    total_row = len(subs) + 2
    dept_list = list(sorted({s["department"] for s in subs}))
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    summary_cell = ws.cell(
        row=total_row,
        column=1,
        value=f"共 {len(subs)} 个部门/班组提交，已提交部门：{'、'.join(dept_list)}",
    )
    summary_cell.font = Font(name="微软雅黑", bold=True, size=10)
    summary_cell.alignment = left_align

    # Missing departments
    if catalog["departments"]:
        authorized = [d.strip() for d in catalog["departments"].split(",") if d.strip()]
    else:
        authorized = list(DEPARTMENTS)
    missing = [d for d in authorized if d not in dept_list]
    if missing:
        missing_row = total_row + 1
        ws.merge_cells(start_row=missing_row, start_column=1, end_row=missing_row, end_column=6)
        missing_cell = ws.cell(
            row=missing_row, column=1, value=f"未提交部门：{'、'.join(missing)}"
        )
        missing_cell.font = Font(name="微软雅黑", bold=True, size=10, color="FF0000")
        missing_cell.alignment = left_align

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 30

    ledger_path = gen_root / "汇总台账.xlsx"
    wb.save(str(ledger_path))

    # Copy files into department folders
    for sub in subs:
        dept_dir = gen_root / sub["department"]
        dept_dir.mkdir(parents=True, exist_ok=True)
        # sign-in file
        if sub["sign_in_file"]:
            src = UPLOAD_DIR / sub["sign_in_file"]
            if src.exists():
                ext = src.suffix
                shutil.copy2(str(src), str(dept_dir / f"签到单{ext}"))
        # photos
        photos = json.loads(sub["photos"] or "[]")
        photos_dir = dept_dir / "照片"
        photos_dir.mkdir(parents=True, exist_ok=True)
        for i, pp in enumerate(photos, 1):
            src = UPLOAD_DIR / pp
            if src.exists():
                ext = src.suffix
                shutil.copy2(str(src), str(photos_dir / f"照片{i}{ext}"))

    # Create ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fpath in gen_root.rglob("*"):
            if fpath.is_file():
                arcname = str(fpath.relative_to(gen_root))
                zf.write(str(fpath), arcname)
    zip_buf.seek(0)

    # Cache for download
    zip_path = GENERATED_DIR / str(catalog_id) / f"{catalog_name}.zip"
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    zip_path.write_bytes(zip_buf.getvalue())

    return jsonify({"ok": True, "filename": f"{catalog_name}.zip"})


@bp.get("/api/collection/download/<int:catalog_id>")
def download(catalog_id):
    data = request.args.to_dict()
    pwd = data.get("password", "")
    if not pwd or pwd != TRAINING_EDIT_PASSWORD:
        return jsonify({"error": "密码错误"}), 403
    filename = data.get("filename", "").strip()
    if not filename:
        abort(400)
    zip_path = GENERATED_DIR / str(catalog_id) / filename
    if not zip_path.exists():
        abort(404)
    return send_file(str(zip_path), as_attachment=True, download_name=filename)


# ── Departments list ──

@bp.get("/api/collection/departments")
def get_departments():
    return jsonify(DEPARTMENTS)
