import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def value(cell):
    result = cell.value
    if result is None or isinstance(result, (str, int, float, bool)):
        return result
    return str(result)


for name in sys.argv[1:]:
    path = Path(name)
    workbook = load_workbook(path, read_only=True, data_only=False)
    print(json.dumps({"file": str(path), "sheets": workbook.sheetnames}, ensure_ascii=False))
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 80), max_col=min(sheet.max_column, 40)):
            values = [value(cell) for cell in row]
            if any(item not in (None, "") for item in values):
                rows.append({"row": row[0].row, "values": values})
        print(json.dumps({
            "sheet": sheet.title,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column,
            "rows": rows,
        }, ensure_ascii=False))
